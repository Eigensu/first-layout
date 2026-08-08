"""Importer mapping for the auction sale value.

The template's Points column has always fed Player.price. The dedicated Price
column added for auction contests takes precedence when present, and older
sheets that only carry Points keep working unchanged.
"""

from app.utils.import_players.import_validators import validate_player_row


async def _row(**fields):
    base = {"name": "Ankit Shah", "team": "DV SPARTANS", "status": "Active"}
    base.update(fields)
    data, error = await validate_player_row(base, slot_strategy="ignore")
    return data, error


async def test_price_column_sets_the_auction_value(db):
    data, error = await _row(points=1000, price=200_000)
    assert error is None
    assert data["price"] == 200_000
    # Accumulated fantasy points always start at zero.
    assert data["points"] == 0


async def test_points_column_still_feeds_price_when_price_absent(db):
    data, error = await _row(points=1000)
    assert error is None
    assert data["price"] == 1000


async def test_blank_price_falls_back_to_points(db):
    data, error = await _row(points=1000, price="  ")
    assert error is None
    assert data["price"] == 1000


async def test_player_with_neither_is_left_unauctioned(db):
    data, error = await _row()
    assert error is None
    assert data["price"] == 0


async def test_price_does_not_leak_into_stats(db):
    data, error = await _row(points=1000, price=200_000, runs=742)
    assert error is None
    stats = data["stats"] or {}
    assert "price" not in stats
    assert stats.get("runs") == 742


async def test_negative_price_is_rejected(db):
    data, error = await _row(price=-5)
    assert error is not None
    assert error.field == "price"


# --- generated templates stay aligned with their own columns ---------------


async def test_csv_example_row_matches_the_column_order(db):
    """
    The generated template must be importable as-is. A row that drifts out of
    step with TEMPLATE_COLUMNS puts text in the price column and shifts every
    field after it.
    """
    import csv
    import io

    from app.utils.import_players.import_template import (
        TEMPLATE_COLUMNS,
        generate_csv_template,
    )
    from app.utils.import_players.import_parsers import normalize_header

    rows = list(csv.reader(io.StringIO(generate_csv_template())))
    header, example = rows[0], rows[1]

    assert header == TEMPLATE_COLUMNS
    assert len(example) == len(TEMPLATE_COLUMNS)

    row = {normalize_header(h): v for h, v in zip(header, example)}
    data, error = await validate_player_row(row, slot_strategy="ignore")

    assert error is None, f"generated example row does not import: {error}"
    assert data["price"] == 200_000
    assert data["team"] == "DV SPARTANS"


async def test_xlsx_example_row_matches_the_column_order(db):
    from openpyxl import load_workbook

    from app.utils.import_players.import_template import (
        TEMPLATE_HEADERS,
        generate_xlsx_template,
    )
    from app.utils.import_players.import_parsers import normalize_header

    ws = load_workbook(await generate_xlsx_template())["Players"]
    header = [c.value for c in ws[1]]
    example = [c.value for c in ws[2]]

    assert header == TEMPLATE_HEADERS
    assert len(example) == len(TEMPLATE_HEADERS)

    row = {normalize_header(str(h)): v for h, v in zip(header, example)}
    data, error = await validate_player_row(row, slot_strategy="ignore")

    assert error is None, f"generated example row does not import: {error}"
    assert data["price"] == 200_000
